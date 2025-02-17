import os
import time
import threading
from flask import Flask, request, render_template, send_file, redirect, url_for, jsonify
from zapv2 import ZAPv2
import openpyxl
from openpyxl.styles import Font

app = Flask(__name__)

# OWASP ZAP API and proxy configuration
ZAP_API_KEY = 'f3fh33gcn9rda93gttg5dg0iad'  # API key added here
ZAP_PROXY = 'http://localhost:8080'
scan_progress = 0  # Global variable to track scan progress

# Function to generate Excel report
def create_excel_report(alerts, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "OWASP ZAP Scan Report"
    headers = ["Alert", "Risk", "Confidence", "Description", "URL", "Solution", "Reference"]
    header_font = Font(bold=True)

    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header).font = header_font

    for row_num, alert in enumerate(alerts, start=2):
        sheet.cell(row=row_num, column=1, value=alert.get('alert'))
        sheet.cell(row=row_num, column=2, value=alert.get('risk'))
        sheet.cell(row=row_num, column=3, value=alert.get('confidence'))
        sheet.cell(row=row_num, column=4, value=alert.get('description'))
        sheet.cell(row=row_num, column=5, value=alert.get('url'))
        sheet.cell(row=row_num, column=6, value=alert.get('solution'))
        sheet.cell(row=row_num, column=7, value=alert.get('reference'))

    workbook.save(output_file)
    print(f"Report saved as {output_file}")

# ZAP Scan Function
def run_zap_scan(target_url, output_file):
    global scan_progress
    zap = ZAPv2(apikey=ZAP_API_KEY, proxies={'http': ZAP_PROXY, 'https': ZAP_PROXY})
    print(f"Starting scan on {target_url}")

    # Reset progress
    scan_progress = 0

    # Spider scan
    print("Starting spider scan...")
    zap.spider.scan(target_url)
    while int(zap.spider.status()) < 100:
        time.sleep(2)
        scan_progress = int(zap.spider.status()) // 2  # 50% progress for spider scan
        print(f"Spider scan progress: {zap.spider.status()}%")
    scan_progress = 50  # Spider scan completed

    # Active scan
    print("Starting active scan...")
    zap.ascan.scan(target_url)
    while int(zap.ascan.status()) < 100:
        time.sleep(2)
        scan_progress = 50 + (int(zap.ascan.status()) // 2)  # Remaining 50% for active scan
        print(f"Active scan progress: {zap.ascan.status()}%")
    scan_progress = 90  # Active scan completed

    # Get alerts and create a report
    alerts = zap.core.alerts()
    create_excel_report(alerts, output_file)

    # Finalize progress
    scan_progress = 100
    print("Scan completed successfully.")

# Flask Routes
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        target_url = request.form.get("target_url")
        output_file = "results.xlsx"

        # Start the scan in a new thread
        thread = threading.Thread(target=run_zap_scan, args=(target_url, output_file))
        thread.start()

        return redirect(url_for("progress"))
    return render_template("index.html")

@app.route("/progress")
def progress():
    return render_template("progress.html")

@app.route("/get_progress")
def get_progress():
    global scan_progress
    return jsonify({"progress": scan_progress})

@app.route("/download")
def download():
    output_file = "results.xlsx"
    if os.path.exists(output_file):
        return send_file(output_file, as_attachment=True)
    return "Report not ready. Try again later.", 404

if __name__ == "__main__":
    app.run(debug=True)
