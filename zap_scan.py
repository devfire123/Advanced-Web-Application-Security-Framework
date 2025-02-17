import time
import openpyxl
from openpyxl.styles import Font, Alignment
from zapv2 import ZAPv2

# Hardcoded OWASP ZAP API keyings
ZAP_API_KEY = 'f3fh33gcn9rda93gttg5dg0iad'
ZAP_PROXY = 'http://127.0.0.1:8080'

def create_excel_report(alerts, output_file):
    # Create a workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "OWASP ZAP Scan Report"

    # Define the headers
    headers = [
        "Alert", "Risk", "Confidence", "Description", "URL", "Solution", "Reference"
    ]
    header_font = Font(bold=True)
    alignment = Alignment(horizontal="left", vertical="center")  # Change alignment to left

    # Write the headers to the first row
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = alignment

    # Write the alerts to the subsequent rows
    for row_num, alert in enumerate(alerts, start=2):
        sheet.cell(row=row_num, column=1, value=alert.get('alert'))
        sheet.cell(row=row_num, column=2, value=alert.get('risk'))
        sheet.cell(row=row_num, column=3, value=alert.get('confidence'))
        sheet.cell(row=row_num, column=4, value=alert.get('description'))
        sheet.cell(row=row_num, column=5, value=alert.get('url'))
        sheet.cell(row=row_num, column=6, value=alert.get('solution'))
        sheet.cell(row=row_num, column=7, value=alert.get('reference'))

    # Adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Save the workbook to the specified file
    workbook.save(output_file)
    print(f'Report generated successfully: {output_file}')

def run_zap_scan(target_url, output_file):
    # Initialize OWASP ZAP instance
    zap = ZAPv2(apikey=ZAP_API_KEY, proxies={'http': ZAP_PROXY, 'https': ZAP_PROXY})

    # Spider the target URL to discover all pages and endpoints
    print(f'Spidering target URL: {target_url}')
    zap.spider.scan(target_url)
    time.sleep(5)  # Delay to allow spidering to start
    while int(zap.spider.status()) < 100:
        print(f'Spider progress: {zap.spider.status()}%')
        time.sleep(5)

    print('Spidering completed.')

    # Perform active scanning to detect vulnerabilities
    print(f'Active scanning target URL: {target_url}')
    zap.ascan.scan(target_url)
    time.sleep(5)  # Delay to allow active scanning to start
    while int(zap.ascan.status()) < 100:
        print(f'Active scan progress: {zap.ascan.status()}%')
        time.sleep(5)

    print('Active scanning completed.')

    # Generate Excel report
    print(f'Generating Excel report: {output_file}')
    alerts = zap.core.alerts()
    create_excel_report(alerts, output_file)

if __name__ == "__main__":
    target_url = input("Enter the target URL to scan: ")
    output_file = input("Enter the output file name for scan results (e.g., results.xlsx): ")

    if not output_file.endswith('.xlsx'):
        print("Error: Output file name must end with .xlsx")
    else:
        run_zap_scan(target_url, output_file)